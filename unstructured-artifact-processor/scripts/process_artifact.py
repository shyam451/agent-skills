#!/usr/bin/env python3
"""
Unstructured Artifact Processor
Main processing script for extracting data from any document type.
"""

import sys
import json
import tempfile
import os
from pathlib import Path
from datetime import datetime

# Type identification
def identify_artifact(filepath):
    """Identify artifact type from extension and magic bytes."""
    try:
        import magic
        mime = magic.from_file(filepath, mime=True)
    except ImportError:
        mime = None
    
    ext = Path(filepath).suffix.lower()
    
    type_map = {
        '.eml': 'email_eml', '.msg': 'email_msg',
        '.docx': 'word', '.doc': 'word_legacy', '.odt': 'odt', '.rtf': 'rtf',
        '.xlsx': 'excel', '.xls': 'excel_legacy', '.xlsm': 'excel_macro',
        '.csv': 'csv', '.tsv': 'tsv',
        '.pdf': 'pdf',
        '.zip': 'archive_zip', '.rar': 'archive_rar', '.7z': 'archive_7z',
        '.tar': 'archive_tar', '.gz': 'archive_gz',
        '.png': 'image', '.jpg': 'image', '.jpeg': 'image',
        '.tiff': 'image', '.tif': 'image', '.bmp': 'image', '.gif': 'image',
        '.mp3': 'audio', '.wav': 'audio', '.flac': 'audio',
        '.mp4': 'video', '.avi': 'video', '.mov': 'video', '.mkv': 'video',
        '.json': 'json', '.xml': 'xml', '.yaml': 'yaml', '.yml': 'yaml',
        '.html': 'html', '.htm': 'html',
    }
    
    return type_map.get(ext, 'unknown'), mime, ext

# Email extraction
def extract_email_eml(filepath):
    """Extract EML file."""
    import email
    from email import policy
    
    with open(filepath, 'rb') as f:
        msg = email.message_from_binary_file(f, policy=policy.default)
    
    result = {
        'type': 'email',
        'format': 'eml',
        'headers': {
            'from': msg['from'],
            'to': msg['to'],
            'cc': msg['cc'],
            'subject': msg['subject'],
            'date': msg['date'],
            'message_id': msg['message-id'],
        },
        'body': {'plain': None, 'html': None},
        'attachments': []
    }
    
    for part in msg.walk():
        ct = part.get_content_type()
        disp = str(part.get('Content-Disposition', ''))
        
        if 'attachment' not in disp:
            if ct == 'text/plain' and not result['body']['plain']:
                try:
                    result['body']['plain'] = part.get_content()
                except:
                    result['body']['plain'] = str(part.get_payload(decode=True))
            elif ct == 'text/html' and not result['body']['html']:
                try:
                    result['body']['html'] = part.get_content()
                except:
                    result['body']['html'] = str(part.get_payload(decode=True))
        
        filename = part.get_filename()
        if filename:
            payload = part.get_payload(decode=True)
            result['attachments'].append({
                'filename': filename,
                'content_type': ct,
                'size': len(payload) if payload else 0,
                '_data': payload
            })
    
    return result

def extract_email_msg(filepath):
    """Extract MSG file."""
    try:
        import extract_msg
        msg = extract_msg.Message(filepath)
        
        result = {
            'type': 'email',
            'format': 'msg',
            'headers': {
                'from': msg.sender,
                'to': msg.to,
                'cc': msg.cc,
                'subject': msg.subject,
                'date': str(msg.date),
            },
            'body': {
                'plain': msg.body,
                'html': msg.htmlBody,
            },
            'attachments': []
        }
        
        for att in msg.attachments:
            result['attachments'].append({
                'filename': att.longFilename or att.shortFilename,
                'size': len(att.data) if att.data else 0,
                '_data': att.data
            })
        
        return result
    except ImportError:
        return {'type': 'email', 'error': 'extract-msg not installed'}

# Document extraction
def extract_word(filepath):
    """Extract DOCX file."""
    import zipfile
    from lxml import etree
    
    ns = {'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'}
    
    with zipfile.ZipFile(filepath) as zf:
        result = {
            'type': 'word_document',
            'format': 'docx',
            'content': {'paragraphs': [], 'tables': []},
            'images': [],
            'embedded_objects': [],
            'metadata': {}
        }
        
        # Extract text
        doc_xml = zf.read('word/document.xml')
        tree = etree.fromstring(doc_xml)
        
        for para in tree.findall('.//w:p', ns):
            text = ''.join(t.text or '' for t in para.findall('.//w:t', ns))
            if text.strip():
                result['content']['paragraphs'].append(text)
        
        # Extract tables
        for table in tree.findall('.//w:tbl', ns):
            rows = []
            for row in table.findall('.//w:tr', ns):
                cells = []
                for cell in row.findall('.//w:tc', ns):
                    cell_text = ''.join(t.text or '' for t in cell.findall('.//w:t', ns))
                    cells.append(cell_text)
                rows.append(cells)
            if rows:
                result['content']['tables'].append({'rows': rows})
        
        # Extract embedded objects
        for name in zf.namelist():
            if name.startswith('word/embeddings/'):
                result['embedded_objects'].append({
                    'name': name.split('/')[-1],
                    'size': zf.getinfo(name).file_size,
                    '_data': zf.read(name)
                })
            elif name.startswith('word/media/'):
                result['images'].append({
                    'name': name.split('/')[-1],
                    'size': zf.getinfo(name).file_size,
                })
        
        # Metadata
        try:
            core_xml = zf.read('docProps/core.xml')
            core_tree = etree.fromstring(core_xml)
            result['metadata']['creator'] = core_tree.findtext('.//{*}creator')
            result['metadata']['title'] = core_tree.findtext('.//{*}title')
        except:
            pass
    
    return result

# Spreadsheet extraction
def extract_excel(filepath):
    """Extract XLSX file."""
    import openpyxl
    import zipfile
    
    result = {
        'type': 'spreadsheet',
        'format': 'xlsx',
        'sheets': [],
        'embedded_objects': [],
        'metadata': {}
    }
    
    wb = openpyxl.load_workbook(filepath, data_only=True)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_data = {
            'name': sheet_name,
            'rows': [],
            'dimensions': ws.dimensions
        }
        
        for row in ws.iter_rows(values_only=True):
            sheet_data['rows'].append(list(row))
        
        result['sheets'].append(sheet_data)
    
    wb.close()
    
    # Check for embedded objects
    with zipfile.ZipFile(filepath) as zf:
        for name in zf.namelist():
            if 'embeddings' in name or 'oleObject' in name:
                result['embedded_objects'].append({
                    'name': name.split('/')[-1],
                    'size': zf.getinfo(name).file_size,
                    '_data': zf.read(name)
                })
    
    return result

def extract_csv(filepath, delimiter=','):
    """Extract CSV/TSV file."""
    import pandas as pd
    
    try:
        import chardet
        with open(filepath, 'rb') as f:
            encoding = chardet.detect(f.read())['encoding']
    except:
        encoding = 'utf-8'
    
    df = pd.read_csv(filepath, encoding=encoding, delimiter=delimiter)
    
    return {
        'type': 'spreadsheet',
        'format': 'csv' if delimiter == ',' else 'tsv',
        'columns': list(df.columns),
        'row_count': len(df),
        'data': df.head(1000).to_dict(orient='records')  # Limit for large files
    }

# PDF extraction
def extract_pdf(filepath):
    """Extract PDF file."""
    import pdfplumber
    
    result = {
        'type': 'pdf',
        'content': {'pages': [], 'tables': []},
        'metadata': {},
        'images': []
    }
    
    with pdfplumber.open(filepath) as pdf:
        result['metadata']['page_count'] = len(pdf.pages)
        
        for i, page in enumerate(pdf.pages):
            page_data = {
                'number': i + 1,
                'text': page.extract_text() or '',
                'tables': []
            }
            
            tables = page.extract_tables()
            for table in tables:
                if table:
                    page_data['tables'].append({
                        'headers': table[0] if table else [],
                        'rows': table[1:] if len(table) > 1 else []
                    })
            
            result['content']['pages'].append(page_data)
    
    return result

# Archive extraction
def extract_archive_zip(filepath, max_depth=5, current_depth=0):
    """Extract ZIP archive."""
    import zipfile
    
    if current_depth >= max_depth:
        return {'type': 'archive', 'error': 'max_depth_exceeded'}
    
    result = {
        'type': 'archive',
        'format': 'zip',
        'contents': [],
        'summary': {'total_files': 0, 'total_size': 0}
    }
    
    with zipfile.ZipFile(filepath) as zf:
        for info in zf.infolist():
            entry = {
                'name': info.filename,
                'size': info.file_size,
                'compressed_size': info.compress_size,
                'is_dir': info.is_dir()
            }
            
            result['summary']['total_files'] += 1
            result['summary']['total_size'] += info.file_size
            
            # Store data for recursive processing
            if not info.is_dir():
                entry['_data'] = zf.read(info)
            
            result['contents'].append(entry)
    
    return result

# Image extraction
def extract_image(filepath):
    """Extract image metadata and OCR."""
    from PIL import Image
    from PIL.ExifTags import TAGS
    
    result = {
        'type': 'image',
        'metadata': {},
        'exif': {},
        'ocr_text': None
    }
    
    with Image.open(filepath) as img:
        result['metadata'] = {
            'format': img.format,
            'mode': img.mode,
            'width': img.width,
            'height': img.height
        }
        
        exif = img._getexif()
        if exif:
            for tag_id, value in exif.items():
                tag = TAGS.get(tag_id, tag_id)
                try:
                    result['exif'][tag] = str(value)
                except:
                    pass
    
    # OCR
    try:
        import pytesseract
        result['ocr_text'] = pytesseract.image_to_string(Image.open(filepath))
    except:
        pass
    
    return result

# Audio/Video extraction
def extract_audio(filepath):
    """Extract audio metadata."""
    try:
        from mutagen import File as MutagenFile
        audio = MutagenFile(filepath)
        
        result = {
            'type': 'audio',
            'metadata': {
                'duration': getattr(audio.info, 'length', None),
                'bitrate': getattr(audio.info, 'bitrate', None),
                'sample_rate': getattr(audio.info, 'sample_rate', None),
            },
            'tags': {}
        }
        
        if audio.tags:
            for key, value in audio.tags.items():
                result['tags'][str(key)] = str(value)
        
        return result
    except ImportError:
        return {'type': 'audio', 'error': 'mutagen not installed'}

def extract_video(filepath):
    """Extract video metadata."""
    import subprocess
    
    result = {
        'type': 'video',
        'metadata': {},
        'streams': []
    }
    
    try:
        probe = subprocess.run([
            'ffprobe', '-v', 'quiet',
            '-print_format', 'json',
            '-show_format', '-show_streams',
            filepath
        ], capture_output=True, text=True, timeout=30)
        
        if probe.returncode == 0:
            data = json.loads(probe.stdout)
            result['metadata'] = data.get('format', {})
            result['streams'] = data.get('streams', [])
    except:
        pass
    
    return result

# Structured data extraction
def extract_json_file(filepath):
    """Extract JSON file."""
    with open(filepath, 'r') as f:
        data = json.load(f)
    
    return {
        'type': 'json',
        'content': data
    }

def extract_xml(filepath):
    """Extract XML file."""
    from lxml import etree
    
    tree = etree.parse(filepath)
    
    def element_to_dict(elem):
        result = {'tag': elem.tag, 'text': elem.text, 'children': []}
        for child in elem:
            result['children'].append(element_to_dict(child))
        return result
    
    return {
        'type': 'xml',
        'content': element_to_dict(tree.getroot())
    }

def extract_html(filepath):
    """Extract HTML file."""
    from lxml import html
    
    with open(filepath, 'rb') as f:
        tree = html.parse(f)
    
    return {
        'type': 'html',
        'title': tree.findtext('.//title'),
        'text': tree.getroot().text_content()
    }

# Generic extraction
def extract_generic(filepath):
    """Generic text extraction."""
    try:
        with open(filepath, 'r', errors='ignore') as f:
            content = f.read()
        return {'type': 'text', 'content': content}
    except:
        return {'type': 'binary', 'error': 'cannot_read'}

# Main processor
def process_artifact(filepath, max_depth=5, current_depth=0):
    """Main processing function."""
    artifact_type, mime, ext = identify_artifact(filepath)
    
    extractors = {
        'email_eml': extract_email_eml,
        'email_msg': extract_email_msg,
        'word': extract_word,
        'excel': extract_excel,
        'csv': lambda f: extract_csv(f, ','),
        'tsv': lambda f: extract_csv(f, '\t'),
        'pdf': extract_pdf,
        'archive_zip': lambda f: extract_archive_zip(f, max_depth, current_depth),
        'image': extract_image,
        'audio': extract_audio,
        'video': extract_video,
        'json': extract_json_file,
        'xml': extract_xml,
        'html': extract_html,
    }
    
    extractor = extractors.get(artifact_type, extract_generic)
    
    try:
        result = extractor(filepath)
    except Exception as e:
        result = {
            'type': artifact_type,
            'error': str(e),
            'error_type': type(e).__name__
        }
    
    # Add common metadata
    result['source_file'] = Path(filepath).name
    result['artifact_type'] = artifact_type
    result['mime_type'] = mime
    result['processing_timestamp'] = datetime.now().isoformat()
    
    # Recursive processing for embedded content
    if current_depth < max_depth:
        result = process_embedded_content(result, max_depth, current_depth)
    
    return result

def process_embedded_content(result, max_depth, current_depth):
    """Process embedded content recursively."""
    for key in ('attachments', 'embedded_objects', 'contents'):
        if key in result and isinstance(result[key], list):
            for item in result[key]:
                if '_data' in item and item['_data']:
                    # Determine extension
                    name = item.get('filename') or item.get('name', 'unknown')
                    ext = Path(name).suffix or '.bin'
                    
                    with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
                        tmp.write(item['_data'])
                        tmp.flush()
                        
                        try:
                            item['extracted_content'] = process_artifact(
                                tmp.name, max_depth, current_depth + 1
                            )
                        except Exception as e:
                            item['extraction_error'] = str(e)
                        finally:
                            os.unlink(tmp.name)
                    
                    # Remove raw data from output
                    del item['_data']
    
    return result

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python process_artifact.py <filepath> [max_depth]")
        sys.exit(1)
    
    filepath = sys.argv[1]
    max_depth = int(sys.argv[2]) if len(sys.argv) > 2 else 5
    
    result = process_artifact(filepath, max_depth)
    print(json.dumps(result, indent=2, default=str))
